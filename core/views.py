import smtplib
import ssl
import openpyxl
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import login, authenticate
from django.shortcuts import render, redirect, get_object_or_404
from .forms import CustomUserCreationForm
from django.core.mail import send_mail
from django.shortcuts import render, redirect
from .forms import QuizUploadForm
from .models import Quiz
from .models import StudentPerformance
from .models import Quiz, Attendance, StudentPerformance
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from .forms import QuizForm
from datetime import date
from django.utils import timezone
from django.contrib import messages
from django.http import HttpResponseRedirect
from django.contrib.auth import logout
from django.core.paginator import Paginator
from core.models import CustomUser
from datetime import timedelta
from django.db.models import Sum
from django.db.models import Count, Max
from django.contrib.auth.forms import UserChangeForm
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Question, Choice
from django.core.files.storage import FileSystemStorage
from django.http import JsonResponse
from openpyxl import load_workbook

def home(request):
    return render(request, 'core/home.html')



@csrf_exempt 
def logout_view(request):
    if request.method == 'POST':
        logout(request)  # Log out the user
        return redirect('home')  # Redirect to the home page
    return redirect('home')  # Redirect even if accessed via GET (as a fallback)

def register(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Registration successful! Redirecting to login page in 3 seconds.")
            return render(request, 'core/register.html', {'form': form})
    else:
        form = CustomUserCreationForm()

    return render(request, 'core/register.html', {'form': form})

def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect(reverse('profile_dashboard'))
    else:
        form = AuthenticationForm()
    return render(request, 'core/login.html', {'form': form})


from django.contrib import messages
from openpyxl import load_workbook

@login_required
def teacher_dashboard(request):
    if request.user.role != 'teacher':
        return redirect('home')

    if request.method == 'POST':
        if 'excel_file' in request.FILES:
            excel_file = request.FILES['excel_file']
            try:
                workbook = load_workbook(excel_file)
                sheet = workbook.active

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    course_id, question_text, choices, correct_answer = row

                    quiz, created = Quiz.objects.get_or_create(course_id=course_id)
                    
                    # Create the Question
                    question = Question.objects.create(
                        quiz=quiz,
                        text=question_text,
                        correct_answer=correct_answer
                    )

                    # Add choices
                    if choices:
                        choices_list = [choice.strip() for choice in choices.split(',')]
                        for choice in choices_list:
                            Choice.objects.create(question=question, text=choice)

                messages.success(request, "Quiz uploaded successfully!")
                return redirect('teacher_dashboard')
            except Exception as e:
                messages.error(request, f"Error uploading quiz: {e}")
                return redirect('teacher_dashboard')

    quizzes = Quiz.objects.prefetch_related('questions').all()
    paginator = Paginator(quizzes, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'core/teacher_dashboard.html', {
        'page_obj': page_obj,
        'username': request.user.username,
        'now': timezone.localtime(),
    })



@login_required
def student_dashboard(request):
    # Ensure only a student can view this page
    if request.user.role != 'student':
        return redirect('home')  # Or a page where non-students are redirected

    # Fetch quizzes that are upcoming (filter by date)
    quizzes = Quiz.objects.filter(upload_date__gte='2024-12-14')  # Filter by current date
    total_quizzes = quizzes.count()
    for quiz in quizzes:
        quiz.questions_list = quiz.questions.all()  # Get the associated questions for each quiz
        quiz.already_attended = False
        if Attendance.objects.filter(quiz=quiz, user=request.user, is_attended=True).exists():
            quiz.already_attended = True  # Mark as attended if True


    # Fetch student performance and attendance data
    performance_data = StudentPerformance.objects.filter(student=request.user)
    attendance_data = Attendance.objects.filter(user=request.user)

    # Calculate total quizzes attended and total score
    total_attended = attendance_data.filter(is_attended=True).count()
    total_score = performance_data.filter(quiz__attendance__is_attended=True).aggregate(Sum('score'))['score__sum'] or 0
    
    # Paginate quizzes
    quizzes_paginator = Paginator(quizzes, 5)  # Show 5 quizzes per page
    page_number = request.GET.get('quiz_page')  # Get the page number for quizzes
    quiz_page_obj = quizzes_paginator.get_page(page_number)

    # Paginate performance data
    performance_paginator = Paginator(performance_data, 5)  # Show 5 performance records per page
    performance_page_number = request.GET.get('performance_page')  # Get the page number for performance data
    performance_page_obj = performance_paginator.get_page(performance_page_number)

    # Paginate attendance data
    attendance_paginator = Paginator(attendance_data, 5)  # Show 5 attendance records per page
    attendance_page_number = request.GET.get('attendance_page')  # Get the page number for attendance data
    attendance_page_obj = attendance_paginator.get_page(attendance_page_number)



    # Pass the user object to the context to display username
    return render(request, 'core/student_dashboard.html', {
        'quizzes': quiz_page_obj,  # Updated quizzes with pagination
        'performance_data': performance_page_obj,  # Updated performance data with pagination
        'attendance_data': attendance_page_obj,
        'total_quizzes': total_quizzes,
        'total_attended': total_attended,
        'total_score': total_score,
        'username': request.user.username,  # Pass username for display
    })



def send_email_to_student(request):
    try:
        # SMTP settings for Gmail
        smtp_server = "smtp.gmail.com"
        smtp_port = 587
        sender_email = "kiransp64@gmail.com"  # Your Gmail address
        receiver_email = "kiransp46@gmail.com"  # Recipient email
        password = "64@psnarik"  # Use app password if 2FA is enabled
        
        # Create a secure SSL context and disable certificate verification
        context = ssl.create_default_context()
        context.check_hostname = False
        context.verify_mode = ssl.CERT_NONE
        
        # Connect to the SMTP server and start TLS (for encryption)
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls(context=context)  # Secure connection
            server.login(sender_email, password)  # Log in to Gmail SMTP
            subject = "Test Email from Django"
            body = "This is a test email sent via Django using smtplib."
            message = f"Subject: {subject}\n\n{body}"
            
            # Send email
            server.sendmail(sender_email, receiver_email, message)
        
        return HttpResponse("Email sent successfully!")
    
    except Exception as e:
        return HttpResponse(f"Failed to send email: {str(e)}")

def home(request):
    return render(request, 'core/home.html')

def send_quiz_reminder():
    # Get all quizzes that are scheduled within the next 24 hours
    upcoming_quizzes = Quiz.objects.filter(upload_date__gte=timezone.now(), upload_date__lt=timezone.now() + timedelta(days=1))

    for quiz in upcoming_quizzes:
        students = CustomUser.objects.filter(role='student')  # Get all students
        for student in students:
            # Sending the reminder email
            subject = f"Reminder: Upcoming Quiz on {quiz.course_id}"
            message = f"Dear {student.username},\n\nThis is a reminder that the quiz for the course {quiz.course_id} is scheduled to take place soon. Please make sure to attend it.\n\nBest regards,\nYour Team"
            from_email = 'kiransp46@gmail.com'
            recipient_list = [student.email]  # Send the email to the student’s email address

            send_mail(subject, message, from_email, recipient_list)

    return "Emails sent successfully!"


def save_student_performance(student, quiz, score):
    performance = StudentPerformance(student=student, quiz=quiz, score=score, date_taken=date.today())
    performance.save()

@login_required
def take_quiz(request, course_id):

    
    # Fetch all quizzes with the same course_id
    quizzes = Quiz.objects.filter(course_id=course_id)
    
    if not quizzes:
        return render(request, 'core/error.html', {'message': 'No quizzes found for this course.'})
    
    if request.method == "POST":
        total_score = 0
        
        # Iterate over all quizzes
        for quiz in quizzes:
            score = 0
            # Compare selected answers with correct answers and calculate score
            for question in quiz.questions.all():
                selected_answer = request.POST.get(f'question_{question.id}')
                if selected_answer == question.correct_answer:
                    score += 1  
            
            # Save student's performance for this quiz
            StudentPerformance.objects.create(
                student=request.user,
                quiz=quiz,
                score=score,
                date_taken=date.today()
            )
            total_score += score
            Attendance.objects.create(user=request.user, quiz=quiz, is_attended=True)
        
        # Redirect to student dashboard after submitting the quiz
        return redirect('student_dashboard')

    return render(request, 'core/quiz_form.html', {'quizzes': quizzes})




@login_required
def profile_dashboard(request):
    user = request.user  # Get the currently logged-in user
    
    if request.method == 'POST':
        form = UserChangeForm(request.POST, request.FILES, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Your profile has been updated successfully.')
            return redirect('profile_dashboard')
    else:
        form = UserChangeForm(instance=user)

    return render(request, 'core/profile_dashboard.html', {
        'form': form,
       'user': user,
    })



def upload_quiz_from_excel(request):
    if request.method == 'POST':
        form = QuizUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']

            try:
                # Load the Excel file
                wb = openpyxl.load_workbook(excel_file)
                sheet = wb.active  # Assuming the first sheet contains the quiz

                # Example structure: Course ID, Question, Correct Answer, Choices
                for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
                    course_id, question_text, correct_answer, choices = row
                    # Create a new Quiz
                    quiz, created = Quiz.objects.get_or_create(course_id=course_id)

                    # Create a Question
                    question = Question.objects.create(
                        quiz=quiz,
                        question_text=question_text,
                        correct_answer=correct_answer
                    )

                    # Create Choices
                    for choice_text in choices.split(','):
                        Choice.objects.create(question=question, choice_text=choice_text.strip())

                messages.success(request, "Quiz uploaded successfully!")
                return redirect('teacher_dashboard')

            except Exception as e:
                messages.error(request, f"Error processing file: {e}")

    else:
        form = QuizUploadForm()

    return render(request, 'core/upload_quiz.html', {'form': form})

